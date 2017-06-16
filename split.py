#!/usr/bin/python3

import sys

from openpyxl import Workbook
from openpyxl import load_workbook


def split_list(items, parts):
    nb = len(items) / float(parts)
    results = []
    last = 0.0
    while last < len(items):
        results.append(items[int(last):int(last + nb)])
        last += nb
    return results


def get_rows(sheet):
    rows = []
    first = True
    for row in sheet.iter_rows(min_row=1):
        line = []
        for cell in row:
            line.append(cell)
        if first:
            first_line = row
            first = False
        rows.append(row)
    return rows, first_line


def save_values(parts, parts_nb):
    for i in range(0, parts_nb + 1):
        print("Saving {} rows to {}".format(len(parts[i]),
                                    "dest_{}.xlsx".format(i)))
        wb = Workbook()
        ws = wb.active
        y = 1
        for row in parts[i]:
            for cell in row:
                ws.cell(row=y, column=cell.col_idx).value = cell.value
            y += 1
        wb.save("dest_{}.xlsx".format(i))


def starter(filename, parts_nb):
    wb = load_workbook(sys.argv[1])
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    print("Loaded {} of {} lines:".format(sys.argv[1], row_count))
    rows, first_line = get_rows(sheet)
    parts = split_list(rows, parts_nb)
    for part in parts:
        part.insert(0, first_line)
    save_values(parts, parts_nb)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Needs a file name and a parts number")
        sys.exit(1)
    print("Loading {}".format(sys.argv[1]))
    starter(sys.argv[1], int(sys.argv[2]))
